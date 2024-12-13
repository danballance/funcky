import time

from openpyxl.styles import PatternFill
from openpyxl.workbook import Workbook

from funcky.named_constants import TICKS_PER_BAR
from funcky.sequences.note_mono_sequence import NoteMonoSequence


class DebugDumper:
    """
    Quick and dirty script for dumping debug data to xlxs file for investigation
    """
    _track_name: str
    _data: dict[str, list[NoteMonoSequence]]

    def __init__(self, track_name: str, data: dict[str, list[NoteMonoSequence]]):
        self._track_name = track_name
        self._data = data

    def _sequence_to_row(self, i: int, key:str, seq: NoteMonoSequence) -> list[str]:
        row = [f"{i} {key}"]
        for note in seq._items:
            if note is None:
                row.append('')
            else:
                # Format: "note|duration|velocity"
                row.append(f"{note.note}|{note.duration}|{note.velocity}")
        return row

    def _make_header_row_ticks(self) -> list[str]:
        row = [""]
        for i in range(1, TICKS_PER_BAR + 1):
            row.append(f"{i}")
        return row

    def _make_header_row_notes(self) -> list[str]:
        note_cols = {
            1: "1/4",
            7: "2/16",
            13: "1/8",
            19: "4/16",
            25: "2/4",
            31: "6/16",
            37: "7/16",
            43: "8/16",
            49: "3/4",
            55: "10/16",
            61: "11/16",
            67: "4/4",
            73: "13/16",
            79: "14/16",
            85: "15/16",
            91: "16/16",
        }
        row = [""]
        for i in range(1, TICKS_PER_BAR + 1):
            if i in note_cols:
                row.append(note_cols[i])
            else:
                row.append("")
        return row

    def _apply_column_formatting(self, sheet) -> None:
        # Define fills
        quarter_note_fill = PatternFill(start_color="999999", end_color="999999", fill_type="solid")  # Dark green
        eighth_note_fill = PatternFill(start_color="afd095", end_color="afd095", fill_type="solid")  # Medium green
        sixteenth_note_fill = PatternFill(start_color="cccccc", end_color="cccccc", fill_type="solid")  # Light Grey

        # Determine how many rows we have
        max_row = sheet.max_row
        max_col = sheet.max_column

        # Apply formatting based on column indices
        # One bar = 96 columns. Adjust logic if needed.
        for col_idx in range(1, max_col + 1):
            # Determine fill based on musical boundaries
            # Note: zero-based indexing for modulo calculations
            col_zero_based = col_idx - 1
            fill = None
            if col_zero_based % 24 == 0:  # Quarter note boundary
                fill = quarter_note_fill
            elif col_zero_based % 12 == 0:  # Eighth note boundary
                fill = eighth_note_fill
            elif col_zero_based % 6 == 0:  # Sixteenth note boundary
                fill = sixteenth_note_fill

            if fill is not None:
                for row_idx in range(1, max_row + 1):
                    cell = sheet.cell(row=row_idx, column=col_idx + 1)
                    cell.fill = fill

    def dump(self) -> None:
        wb = Workbook()
        # Create a "combined" sheet
        combined_sheet = wb.active
        combined_sheet.title = "Combined"
        # For each key, create a new sheet and write sequences
        header_row_ticks = self._make_header_row_ticks()
        header_row_notes = self._make_header_row_notes()
        all_sequences = {}
        for key, sequences in self._data.items():
            all_sequences[key] = []
            sheet = wb.create_sheet(title=key)
            sheet.append(header_row_ticks)
            sheet.append(header_row_notes)
            i = 1
            for seq in sequences:
                row = self._sequence_to_row(i, key, seq)
                sheet.append(row)
                all_sequences[key].append(row)
                i+= 1
            self._apply_column_formatting(sheet)
        # Write all combined sequences
        combined_sheet.append(header_row_ticks)
        combined_sheet.append(header_row_notes)
        keys = list(all_sequences.keys())
        for i in range(0, len(all_sequences[keys[0]])):
            for key in keys:
                combined_sheet.append(all_sequences[key][i])
        # Apply column formatting to the combined sheet
        self._apply_column_formatting(combined_sheet)
        # Save the workbook
        wb.save(f"debug.{self._track_name}.{int(time.time())}.xlsx")
